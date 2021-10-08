import os
import pandas as pd
import json
import openpyxl as openpyxl
import config.settings as settings
heatmaps = {}

def parse_heatmaps(heatmap_template_questions):
    file = 'data/heatmaps/wind.xlsx'
    folder = 'data/heatmapsV2/'
    files = os.listdir(folder)
    filenames = []

    for filename in files:
        if '.xlsx' in filename:
            filenames.append(filename)

    if os.path.isfile('data/all_heatmaps.json'):
        with open("data/all_heatmaps.json", "r") as jsonfile:
            global heatmaps
            heatmaps = json.load(jsonfile)

    else:
        for filename in filenames:
            parse_heatmap(folder + filename, heatmap_template_questions)

        with open("data/all_heatmaps.json", "w") as outfile:
            json.dump(heatmaps, outfile)

    return heatmaps

def parse_heatmap(file, heatmap_template_questions):
    wb = openpyxl.load_workbook(filename=file)
    first_tab = wb[wb.sheetnames[0]]

    heatmap_name = first_tab['B1'].value
    if isinstance(heatmap_name, str):
        heatmap_name = heatmap_name.strip()
    else:
        heatmap_name = ''

    parsed_answers = {}

    answers_tab = wb[wb.sheetnames[1]]
    count = 0


    for row in answers_tab.iter_rows(max_col=10, min_row=5):
        count += 1
        heatmap_question_code = row[2].value
        actual_question_code = heatmap_template_questions['conversion_lookup'].get(heatmap_question_code)

        if actual_question_code and  actual_question_code not in parsed_answers:
            answer = row[7].value == 'Yes'
            notes =  row[8].value
            parsed_answers[actual_question_code] = {
                'value': answer,
                'notes': notes
            }

    heatmaps[heatmap_name] = parsed_answers



def new_business_activity_questions():
    # file1 = '20201216 Business Activity Heatmap Template V0.2.xlsx'
    file = 'data/new/20210226 HSBC heatmap taxonomy.xlsx'

    dead_heatmaps = []
    wb = openpyxl.load_workbook(filename=file)
    first_tab = wb[wb.sheetnames[0]]
    questions = {}

    q1 = ''
    q2 = ''
    q3 = ''
    # q1code = ''
    # q2code = ''
    # q3code = ''
    q1count = 0
    q2count = 0
    q3count = 0


    for row in first_tab.iter_rows(max_col=12, min_row=3, max_row=118):
        sector_activity = row[1].value
        sub_sector_activity = row[2].value
        category = row[11].value
        heatmap = row[11].value
        if isinstance(heatmap, str):
            heatmap = heatmap.strip()
        else:
            heatmap = ''



        if sector_activity:
            if category != q3:
                q3 = category

            # REMNAMT OF OLD REQUIREMENTS - REFACtOR CODES OUT MAYBE
            q1_code = str(q1count)
            q2_code = "{}.{}".format(q1count, q2count)
            q3_code = "{}.{}.{}".format(q1count, q2count, q3count)

            if sector_activity != q1:
                q1 = sector_activity
                q1count += 1
                q2count = 0
                q3count = 0
                q1_code = str(q1count)

                questions[q1_code] = {
                    'code': q1_code,
                    'text': q1
                }



            if sub_sector_activity != q2:
                q2 = sub_sector_activity
                q2count += 1
                q3count = 0
                q2_code = "{}.{}".format(q1count, q2count)
                questions[q1_code][q2_code] = {
                    'code': q2_code,
                    'text': q2
                }


            unavailable = heatmap not in heatmaps
            if unavailable:
                dead_heatmaps.append(heatmap)

            if heatmap:

                if q3 not in questions[q1_code][q2_code]:
                    if q3 == '' or q3 == 'See sub-sector activity':
                        questions[q1_code][q2_code]['heatmap'] = heatmap
                        questions[q1_code][q2_code]['unavailable'] = unavailable

                    else:
                        q3count += 1
                        q3_code = "{}.{}.{}".format(q1count, q2count, q3count)
                        questions[q1_code][q2_code][q3_code] = {
                            'code': q3_code,
                            'text': q3,
                            'heatmap': heatmap,
                            'unavailable': unavailable,
                        }

    return questions


def parse_heatmap_template():
    folder = 'data/new/'
    file = '20210225 BA & Company characteristics V0.2_for UNGC workshop.xlsx'
    # file = '20210215 BA & Company characteristics V0.2.xlsx'
    file = folder + file
    wb = openpyxl.load_workbook(filename=file)
    first_tab = wb[wb.sheetnames[0]]

    # Heatmap code -> Question layout code
    conversion_lookup = {}

    layout = {}
    questions = {}
    curren_be_code = ''
    question_count = 1

    for row in first_tab.iter_rows(max_col=11, min_row=5):
        be_code = row[0].value
        be_title = row[1].value
        question_code = row[7].value
        heatmap_question_code = row[2].value
        question_risk = row[8].value.capitalize()
        question_text = row[10].value
        question_example = row[5].value

        if be_code and be_code not in layout:
            current_be_code = be_code
            question_count = 0
            layout[be_code] = {
                'title': be_title,
                'questions': []
            }

        if question_code:
            question_count += 1
            duplicate = question_code in questions
            if not duplicate:
                questions[question_code] = {
                    'text': question_text,
                    'risk': question_risk,
                    'heatmap_question_code': heatmap_question_code, # going to try and not use this anywhere tho
                    'example': question_example,
                    'be': current_be_code,
                    'location': "{} Q{}".format(current_be_code, question_count),
                }

            layout[current_be_code]['questions'].append({
                'code': question_code,
                'number': question_count,
                'duplicate': duplicate,
            })

            if heatmap_question_code:
                conversion_lookup[heatmap_question_code] = question_code


    return {
        'layout': layout,
        'questions': questions,
        'conversion_lookup': conversion_lookup
    }
